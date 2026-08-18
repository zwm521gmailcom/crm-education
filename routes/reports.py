"""报表:消课数据汇总等。"""
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from extensions import db
from models import Course, Enrollment, Schedule, ScheduleAttendance

bp = Blueprint("reports", __name__, url_prefix="/reports")

# 粒度选项
GRANULARITIES = (
    ("day", "按日"),
    ("week", "按周"),
    ("month", "按月"),
)


def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _bucket_key(d, granularity):
    """根据粒度计算聚合键(yyyy-mm-dd 字符串,用于分组)。"""
    if granularity == "week":
        monday = d - timedelta(days=d.weekday())
        return monday.strftime("%Y-%m-%d")
    if granularity == "month":
        return d.strftime("%Y-%m-01")
    return d.strftime("%Y-%m-%d")  # day


def _bucket_label(key, granularity):
    """从 bucket_key 反推展示用 label。"""
    if granularity == "week":
        monday = datetime.strptime(key, "%Y-%m-%d").date()
        sunday = monday + timedelta(days=6)
        return f"{monday.strftime('%Y-%m-%d')} ~ {sunday.strftime('%m-%d')}"
    if granularity == "month":
        return key[:7]  # yyyy-mm
    return key  # yyyy-mm-dd


def _build_groups(start_str, end_str, granularity):
    """执行 SQL 查询 + Python 聚合,返回 (rows, totals) 元组。

    rows: [{bucket, label, course, teacher, sessions, students, hours, value, avg_price}, ...]
    totals: {sessions, students, hours, value}  # students 这里是 sum-of-distinct(每行去重数加总)
    """
    q = (
        db.session.query(
            Schedule.start_time,
            Course.name.label("course_name"),
            Schedule.teacher,
            ScheduleAttendance.id.label("att_id"),
            ScheduleAttendance.enrollment_id,
            ScheduleAttendance.hours_used,
            Enrollment.unit_price,
        )
        .join(ScheduleAttendance, ScheduleAttendance.schedule_id == Schedule.id)
        .join(Enrollment, ScheduleAttendance.enrollment_id == Enrollment.id)
        .join(Course, Schedule.course_id == Course.id)
        .filter(ScheduleAttendance.attendance.in_(["present", "makeup"]))
        .filter(ScheduleAttendance.hours_used > 0)
        .filter(Enrollment.status != "refunded")
    )

    start = _parse_date(start_str)
    end = _parse_date(end_str)
    if start:
        q = q.filter(Schedule.start_time >= datetime.combine(start, datetime.min.time()))
    if end:
        q = q.filter(Schedule.start_time < datetime.combine(end + timedelta(days=1), datetime.min.time()))

    raw = q.order_by(Schedule.start_time.asc()).all()

    groups = defaultdict(lambda: {
        "sessions": 0,
        "students": set(),
        "hours": 0.0,
        "value": 0.0,
    })
    for r in raw:
        if not r.start_time:
            continue
        d = r.start_time.date()
        bk = _bucket_key(d, granularity)
        teacher = (r.teacher or "").strip() or "(未指定)"
        g = groups[(bk, r.course_name, teacher)]
        g["sessions"] += 1
        g["students"].add(r.enrollment_id)
        g["hours"] += float(r.hours_used or 0)
        g["value"] += float(r.hours_used or 0) * float(r.unit_price or 0)

    rows = []
    for (bk, course, teacher), g in groups.items():
        hours = g["hours"]
        value = g["value"]
        rows.append({
            "bucket": bk,
            "label": _bucket_label(bk, granularity),
            "course": course,
            "teacher": teacher,
            "sessions": g["sessions"],
            "students": len(g["students"]),
            "hours": hours,
            "value": value,
            "avg_price": (value / hours) if hours > 0 else 0.0,
        })

    # 默认按 bucket 倒序(最新在前),bucket 内按课时倒序
    rows.sort(key=lambda x: (x["bucket"], x["course"], x["teacher"]), reverse=True)

    totals = {
        "sessions": sum(r["sessions"] for r in rows),
        "students": sum(r["students"] for r in rows),  # 各组学员数累加
        "hours": sum(r["hours"] for r in rows),
        "value": sum(r["value"] for r in rows),
    }
    return rows, totals


@bp.route("/consumed")
def consumed():
    granularity = request.args.get("granularity", "day")
    if granularity not in ("day", "week", "month"):
        granularity = "day"
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()

    rows, totals = _build_groups(start_str, end_str, granularity)

    return render_template(
        "reports/consumed.html",
        granularity=granularity,
        granularities=GRANULARITIES,
        rows=rows,
        totals=totals,
        start=start_str,
        end=end_str,
    )


# Excel 导出样式
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


@bp.route("/consumed.xlsx")
def consumed_xlsx():
    granularity = request.args.get("granularity", "day")
    if granularity not in ("day", "week", "month"):
        granularity = "day"
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()

    rows, totals = _build_groups(start_str, end_str, granularity)

    gname = {"day": "日", "week": "周", "month": "月"}[granularity]
    headers = [
        f"周期({gname})", "课程名称", "老师",
        "上课人次", "学员数", "消耗课时", "消耗金额", "平均单价(元/小时)",
    ]
    data = []
    for r in rows:
        data.append([
            r["label"], r["course"], r["teacher"],
            r["sessions"], r["students"],
            round(r["hours"], 2), round(r["value"], 2),
            round(r["avg_price"], 2),
        ])
    # 合计行
    total_hours = totals["hours"]
    data.append([
        "合计", "", "",
        totals["sessions"], totals["students"],
        round(total_hours, 2), round(totals["value"], 2),
        round((totals["value"] / total_hours) if total_hours > 0 else 0, 2),
    ])

    wb = Workbook()
    ws = wb.active
    ws.title = f"消课汇总({gname})"
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    for row_idx, row in enumerate(data, 2):
        for col_idx, v in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = LEFT
    # 列宽
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in data:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else None
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"消课汇总_{gname}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
