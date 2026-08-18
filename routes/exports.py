"""Excel 导出:学员/报名/收款/退费/排期 5 张表。"""
from datetime import datetime
from io import BytesIO

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from extensions import db
from models import (
    Course,
    Enrollment,
    Payment,
    Refund,
    Schedule,
    ScheduleAttendance,
    Student,
)
from routes.students import _apply_student_sort, STUDENT_SORT_KEYS

bp = Blueprint("exports", __name__)


# 通用样式
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _autosize(ws, headers, rows):
    """根据内容自动调列宽。"""
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows:
            v = row[col_idx - 1] if col_idx - 1 < len(row) else None
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def _write_workbook(headers, rows, sheet_name="数据"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 表头
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    # 数据
    for row_idx, row in enumerate(rows, 2):
        for col_idx, v in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = LEFT
    _autosize(ws, headers, rows)
    # 冻结表头
    ws.freeze_panes = "A2"
    return wb


def _send(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _ts():
    return datetime.now().strftime("%Y%m%d_%H%M")


@bp.route("/students.xlsx")
def students_xlsx():
    # 排序参数(与学员管理页共用)
    sort = request.args.get("sort", "created_at")
    if sort not in STUDENT_SORT_KEYS:
        sort = "created_at"
    direction = request.args.get("dir", "desc")
    if direction not in ("asc", "desc"):
        direction = "desc"

    # 按学员聚合 4 个原始数(排除已退费)
    agg_rows = (
        db.session.query(
            Enrollment.student_id,
            func.coalesce(func.sum(Enrollment.total_hours), 0),       # 报课总节数
            func.coalesce(func.sum(Enrollment.used_hours), 0),        # 累计消课节数
            func.coalesce(func.sum(Enrollment.final_price), 0),        # 预收总金额
            func.coalesce(func.sum(Enrollment.used_hours * Enrollment.unit_price), 0),  # 累计消课金额
        )
        .filter(Enrollment.status != "refunded")
        .group_by(Enrollment.student_id)
        .all()
    )
    agg_map = {}
    for sid, total_h, used_h, total_v, used_v in agg_rows:
        used_v = float(used_v)
        used_h = float(used_h)
        avg_price = (used_v / used_h) if used_h > 0 else 0.0
        remaining_h = max(0.0, float(total_h) - used_h)
        remaining_v = max(0.0, float(total_v) - used_v)
        agg_map[sid] = {
            "total_hours": float(total_h),
            "used_hours": used_h,
            "total_value": float(total_v),
            "used_value": used_v,
            "avg_price": avg_price,
            "remaining_hours": remaining_h,
            "remaining_value": remaining_v,
        }

    query = _apply_student_sort(Student.query, sort, direction)
    rows = []
    for s in query.all():
        primary = s.primary_contact
        a = agg_map.get(s.id, {
            "total_hours": 0, "used_hours": 0, "total_value": 0,
            "used_value": 0, "avg_price": 0,
            "remaining_hours": 0, "remaining_value": 0,
        })
        rows.append([
            s.code,
            s.name,
            {"male": "男", "female": "女", "unknown": "-"}.get(s.gender, "-"),
            s.birth_date.isoformat() if s.birth_date else "",
            s.school or "",
            s.grade or "",
            s.phone or "",
            primary.phone if primary else "",
            s.address or "",
            s.enrollment_date.isoformat() if s.enrollment_date else "",
            s.status_label,
            s.total_remaining_hours,
            a["total_hours"],
            a["avg_price"],
            a["total_value"],
            a["used_hours"],
            a["used_value"],
            a["remaining_hours"],
            a["remaining_value"],
            s.notes or "",
        ])
    headers = [
        "编号", "姓名", "性别", "出生日期", "学校", "年级",
        "本人电话", "主联系人电话", "地址", "入学日期", "状态",
        "剩余课时",
        "报课总节数", "课时单价(元/小时)", "预收总金额",
        "累计消课节数", "累计消课金额",
        "剩余课时节数", "剩余课时金额",
        "备注",
    ]
    wb = _write_workbook(headers, rows, "学员")
    return _send(wb, f"学员_{_ts()}.xlsx")


@bp.route("/enrollments.xlsx")
def enrollments_xlsx():
    rows = []
    for e in Enrollment.query.order_by(Enrollment.id.desc()).all():
        rows.append([
            e.code,
            e.student.name,
            e.course.name,
            e.total_hours,
            e.used_hours,
            e.remaining_hours,
            float(e.unit_price or 0),
            float(e.total_price or 0),
            float(e.discount or 0),
            float(e.final_price or 0),
            float(e.paid_amount),
            float(e.refunded_amount),
            float(e.outstanding),
            e.status_label,
            e.enrolled_at.isoformat() if e.enrolled_at else "",
            e.expires_at.isoformat() if e.expires_at else "",
            e.notes or "",
        ])
    headers = [
        "订单号", "学员", "课程", "总课时", "已用", "剩余",
        "单价", "应付", "优惠", "实付", "已收", "已退", "未结",
        "状态", "报名日期", "有效期至", "备注",
    ]
    wb = _write_workbook(headers, rows, "报名")
    return _send(wb, f"报名_{_ts()}.xlsx")


@bp.route("/payments.xlsx")
def payments_xlsx():
    rows = []
    for p in Payment.query.order_by(Payment.paid_at.desc()).all():
        rows.append([
            p.code,
            p.student.name,
            p.enrollment.code if p.enrollment else "",
            float(p.amount),
            p.type_label,
            p.method_label,
            p.paid_at.strftime("%Y-%m-%d %H:%M") if p.paid_at else "",
            p.notes or "",
        ])
    headers = ["收据号", "学员", "关联报名", "金额", "类型", "方式", "时间", "备注"]
    wb = _write_workbook(headers, rows, "收款")
    return _send(wb, f"收款_{_ts()}.xlsx")


@bp.route("/refunds.xlsx")
def refunds_xlsx():
    rows = []
    for r in Refund.query.order_by(Refund.refunded_at.desc()).all():
        rows.append([
            r.code,
            r.student.name,
            r.enrollment.code if r.enrollment else "",
            r.enrollment.course.name if r.enrollment else "",
            float(r.amount),
            r.method_label,
            r.reason or "",
            r.refunded_at.strftime("%Y-%m-%d %H:%M") if r.refunded_at else "",
            r.notes or "",
        ])
    headers = ["单号", "学员", "关联报名", "课程", "金额", "方式", "原因", "时间", "备注"]
    wb = _write_workbook(headers, rows, "退费")
    return _send(wb, f"退费_{_ts()}.xlsx")


# 排期出勤状态 → 中文
_ATT_LABEL = {
    "present": "出勤",
    "absent": "缺席",
    "makeup": "补课",
    "leave": "请假",
}


@bp.route("/schedules.xlsx")
def schedules_xlsx():
    """排期导出:?view=course 课程视角, ?view=student 学员角度(出勤明细)。"""
    view = request.args.get("view", "course")
    range_ = request.args.get("range", "all")
    course_id = request.args.get("course_id", type=int)
    status = request.args.get("status", "").strip()
    student_id = request.args.get("student_id", type=int)
    attendance = request.args.get("attendance", "").strip()

    if view == "student":
        # 学员角度:每行一条 (排期 × 出勤) 明细
        q = (
            db.session.query(Schedule, ScheduleAttendance, Enrollment, Student)
            .join(ScheduleAttendance, ScheduleAttendance.schedule_id == Schedule.id)
            .join(Enrollment, ScheduleAttendance.enrollment_id == Enrollment.id)
            .join(Student, Enrollment.student_id == Student.id)
        )
        if course_id:
            q = q.filter(Schedule.course_id == course_id)
        if status:
            q = q.filter(Schedule.status == status)
        if student_id:
            q = q.filter(Student.id == student_id)
        if attendance:
            q = q.filter(ScheduleAttendance.attendance == attendance)
        from datetime import datetime as _dt
        if range_ == "upcoming":
            q = q.filter(Schedule.start_time >= _dt.now())
        elif range_ == "past":
            q = q.filter(Schedule.start_time < _dt.now())

        rows = []
        for s, a, e, st in q.order_by(Schedule.start_time.desc()).limit(5000).all():
            rows.append([
                s.start_time.strftime("%Y-%m-%d %H:%M") if s.start_time else "",
                s.end_time.strftime("%Y-%m-%d %H:%M") if s.end_time else "",
                st.name,
                s.course.name,
                s.teacher or "",
                s.classroom or "",
                float(s.hours_per_session or 0),
                _ATT_LABEL.get(a.attendance, a.attendance),
                float(a.hours_used or 0),
                s.status_label,
                e.code,
                a.notes or "",
            ])
        headers = [
            "开始", "结束", "学员", "课程", "老师", "教室",
            "课时/节", "出勤", "扣课时", "排期状态", "关联报名", "备注",
        ]
        wb = _write_workbook(headers, rows, "出勤明细")
        return _send(wb, f"出勤明细_{_ts()}.xlsx")

    # 课程视角
    q = Schedule.query
    if course_id:
        q = q.filter(Schedule.course_id == course_id)
    if status:
        q = q.filter(Schedule.status == status)
    from datetime import datetime as _dt
    if range_ == "upcoming":
        q = q.filter(Schedule.start_time >= _dt.now())
    elif range_ == "past":
        q = q.filter(Schedule.start_time < _dt.now())

    rows = []
    for s in q.order_by(Schedule.start_time.desc()).limit(2000).all():
        rows.append([
            s.start_time.strftime("%Y-%m-%d %H:%M") if s.start_time else "",
            s.end_time.strftime("%Y-%m-%d %H:%M") if s.end_time else "",
            s.course.name,
            s.teacher or "",
            s.classroom or "",
            float(s.hours_per_session or 0),
            s.max_students,
            s.attended_count,
            s.status_label,
            s.notes or "",
        ])
    headers = [
        "开始", "结束", "课程", "老师", "教室",
        "课时/节", "上限", "已上", "状态", "备注",
    ]
    wb = _write_workbook(headers, rows, "排期")
    return _send(wb, f"排期_{_ts()}.xlsx")
